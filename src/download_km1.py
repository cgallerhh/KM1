from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path
from urllib.parse import quote

import requests
from pypdf import PdfReader
from openpyxl import load_workbook

from km1_common import (
    BASE_URL,
    PROCESSED_DIR,
    RAW_DIR,
    Period,
    candidate_urls,
    ensure_dirs,
    latest_candidate_periods,
    now_iso,
    sha256_file,
    write_json,
)


def encoded_url(filename: str) -> str:
    return BASE_URL + quote(filename)


def probe_url(url: str, timeout: int = 20) -> bool:
    try:
        response = requests.head(url, allow_redirects=True, timeout=timeout)
        if response.status_code == 200:
            return True
        if response.status_code in {403, 405}:
            response = requests.get(url, stream=True, timeout=timeout)
            return response.status_code == 200
    except requests.RequestException:
        return False
    return False


def find_latest(today: date | None = None) -> tuple[Period, str, str]:
    for period in latest_candidate_periods(today=today, max_back=6):
        for filename, _ in candidate_urls(period):
            url = encoded_url(filename)
            if probe_url(url):
                extension = Path(filename).suffix.lstrip(".")
                canonical = f"KM1_Januar_bis_{period.month_name}_{period.year}.{extension}"
                return period, canonical, url
    raise RuntimeError("Keine KM1-Datei in den letzten 6 Rueckwaertsversuchen gefunden.")


def download_file(url: str, output_path: Path) -> None:
    response = requests.get(url, timeout=90)
    response.raise_for_status()
    suffix = output_path.suffix.lower()
    if suffix == ".pdf" and not response.content.startswith(b"%PDF"):
        raise RuntimeError(f"Download ist keine PDF-Datei: {url}")
    if suffix == ".xlsx" and not response.content.startswith(b"PK"):
        raise RuntimeError(f"Download ist keine XLSX-Datei: {url}")
    output_path.write_bytes(response.content)


def main() -> int:
    parser = argparse.ArgumentParser(description="Laedt die aktuellste KM1-PDF-Datei.")
    parser.add_argument("--year", type=int, help="Berichtsjahr erzwingen")
    parser.add_argument("--month", type=int, help="Berichtsmonat erzwingen, 1-12")
    args = parser.parse_args()

    ensure_dirs()

    if args.year and args.month:
        period = Period(args.year, args.month)
        found = None
        for filename_candidate, _ in candidate_urls(period):
            url_candidate = encoded_url(filename_candidate)
            if probe_url(url_candidate):
                found = (filename_candidate, url_candidate)
                break
        if not found:
            raise SystemExit(f"Datei nicht gefunden fuer {period.label}")
        filename, url = found
    else:
        period, filename, url = find_latest()

    output_path = RAW_DIR / filename
    if not output_path.exists():
        download_file(url, output_path)

    stand = None
    if output_path.suffix.lower() == ".pdf":
        reader = PdfReader(str(output_path))
        seitenzahl = len(reader.pages)
        first_page_text = reader.pages[0].extract_text() or ""
        for line in first_page_text.splitlines():
            if line.strip().lower().startswith("stand:"):
                stand = line.strip().split(":", 1)[1].strip()
                break
    else:
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        seitenzahl = len(workbook.sheetnames)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=12, values_only=True):
                for value in row:
                    if isinstance(value, str) and value.strip().lower().startswith("stand:"):
                        stand = value.strip().split(":", 1)[1].strip()
                        break
                if stand:
                    break
            if stand:
                break
        workbook.close()
    metadata = {
        "quell_url": url,
        "dateiname": filename,
        "download_datum": now_iso(),
        "berichtszeitraum": f"Januar bis {period.month_name} {period.year}",
        "stand_laut_deckblatt": stand,
        "seitenzahl": seitenzahl,
        "hash_sha256": sha256_file(output_path),
        "erkannter_neuester_monat": period.month_name,
        "jahr": period.year,
        "monat": period.month,
        "lokaler_pfad": str(output_path.relative_to(Path.cwd())),
    }
    write_json(PROCESSED_DIR / "km1_metadata.json", metadata)
    print(f"OK: {filename} gespeichert ({metadata['seitenzahl']} Seiten).")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"FEHLER beim KM1-Download: {exc}", file=sys.stderr)
        raise SystemExit(1)

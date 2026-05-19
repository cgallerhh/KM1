from __future__ import annotations

import csv
import sys
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

from km1_common import PROCESSED_DIR, ROOT, ensure_dirs, read_json, write_json


def latest_pdf_from_metadata() -> Path:
    metadata = read_json(PROCESSED_DIR / "km1_metadata.json", {})
    if not metadata.get("lokaler_pfad"):
        raise RuntimeError("Keine Metadaten gefunden. Bitte zuerst src/download_km1.py ausfuehren.")
    return ROOT / metadata["lokaler_pfad"]


def extract_xlsx(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    tables_out: list[dict] = []
    rows_out: list[dict] = []
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
        sheet_rows: list[list[str]] = []
        text_lines: list[str] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            while values and values[-1] == "":
                values.pop()
            if not values:
                continue
            sheet_rows.append(values)
            text_lines.append(" ".join(values))
            rows_out.append(
                {
                    "quelle_seite": sheet_index,
                    "tabelle": 1,
                    "zeile": row_index,
                    "werte": values,
                }
            )
        tables_out.append(
            {
                "page": sheet_index,
                "sheet_name": sheet.title,
                "text": "\n".join(text_lines),
                "tables": [sheet_rows],
                "source_type": "xlsx",
            }
        )
    workbook.close()
    return tables_out, rows_out


def extract_pdf(pdf_path: Path) -> tuple[list[dict], list[dict]]:
    tables_out: list[dict] = []
    rows_out: list[dict] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            tables = page.extract_tables() or []
            tables_out.append({"page": page_index, "text": text, "tables": tables})
            for table_index, table in enumerate(tables, start=1):
                for row_index, row in enumerate(table, start=1):
                    row_values = ["" if cell is None else str(cell).replace("\n", " ").strip() for cell in row]
                    rows_out.append(
                        {
                            "quelle_seite": page_index,
                            "tabelle": table_index,
                            "zeile": row_index,
                            "werte": row_values,
                        }
                    )
            if not tables:
                for row_index, line in enumerate(text.splitlines(), start=1):
                    cleaned = line.strip()
                    if cleaned:
                        rows_out.append(
                            {
                                "quelle_seite": page_index,
                                "tabelle": 0,
                                "zeile": row_index,
                                "werte": [cleaned],
                            }
                        )
    return tables_out, rows_out


def write_csv(rows: list[dict], path: Path) -> None:
    max_cols = max((len(row["werte"]) for row in rows), default=0)
    with path.open("w", encoding="utf-8", newline="") as handle:
        fieldnames = ["quelle_seite", "tabelle", "zeile"] + [f"spalte_{i}" for i in range(1, max_cols + 1)]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            data = {key: row[key] for key in ("quelle_seite", "tabelle", "zeile")}
            data.update({f"spalte_{i}": value for i, value in enumerate(row["werte"], start=1)})
            writer.writerow(data)


def main() -> int:
    ensure_dirs()
    pdf_path = latest_pdf_from_metadata()
    if not pdf_path.exists():
        raise RuntimeError(f"Quelldatei nicht gefunden: {pdf_path}")
    if pdf_path.suffix.lower() == ".xlsx":
        tables, rows = extract_xlsx(pdf_path)
    else:
        tables, rows = extract_pdf(pdf_path)
    write_json(PROCESSED_DIR / "km1_raw_tables.json", tables)
    write_csv(rows, PROCESSED_DIR / "km1_raw_tables.csv")
    print(f"OK: {len(rows)} Tabellenzeilen aus {pdf_path.name} extrahiert.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"FEHLER bei KM1-Extraktion: {exc}", file=sys.stderr)
        raise SystemExit(1)
